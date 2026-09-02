"""한국공학대 TIP 학생식당 + E동 레스토랑 주간 식단표 수집·파싱.

흐름:
  1) GET https://ibook.tukorea.ac.kr/viewer/menu02 → 페이지 HTML에서 bookcode 추출
  2) POST https://ibook.tukorea.ac.kr/web/RawFileList (key, bookcode, base64) → XML
     → <file file_url="…xlsx"/> 추출
  3) GET file_url → xlsx 바이너리
  4) openpyxl 파싱 → JSON

캐시 키 `cafeteria:menu` TTL 1시간. APScheduler가 매일 07~14시 매시 정각에 강제 갱신.
TTL이 짧으므로 cron 1회 누락되어도 다음 요청이 cache-aside로 자가 회복한다.
"""
import logging
import re
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

import openpyxl

from app.core.cache import get_cached_json, set_cached_json
from app.core.http_client import get_http_client

logger = logging.getLogger(__name__)

KST = ZoneInfo("Asia/Seoul")
CACHE_KEY = "cafeteria:menu"
CACHE_TTL = 3600

VIEWER_URL = "https://ibook.tukorea.ac.kr/viewer/menu02"
RAWFILE_URL = "https://ibook.tukorea.ac.kr/web/RawFileList"
WEB_KEY = "kpu"

_BOOKCODE_RE = re.compile(r"bookcode\s*=\s*['\"]([A-Z0-9]+)['\"]")
_FILE_URL_RE = re.compile(r'file_url="([^"]+\.xlsx?[^"]*)"')
_FILE_NAME_RE = re.compile(r'<file\s+name="([^"]+)"')

# 날짜 헤더는 보통 "31일"이지만, 주가 달을 넘기면 그 첫 칸만 "9/1일"처럼
# 월을 앞에 붙여 온다(2026-08-31 주차 실측). 월 표기를 옵션으로 받지 않으면
# 그 칸이 통째로 누락돼 해당 요일 식단이 사라진다.
_DATE_HEADER_RE = re.compile(r"^(?:(\d{1,2})\s*[/.]\s*)?(\d{1,2})\s*일")
_MEAL_HEADER_RE = re.compile(r"^(천원의\s*아침밥|중식|석식|조식|간식)")
_TIME_RE = re.compile(r"(\d{1,2}:\d{2})\s*[\n~\-]+\s*(\d{1,2}:\d{2})")
_TITLE_TIP_RE = re.compile(r"TIP\s*학생식당")
_TITLE_E_RE = re.compile(r"E동\s*레스토랑")

# "①코너" 처럼 원문자로 시작해 "코너"로 끝나는 셀 — 그 자체는 메뉴가 아니라
# 바로 아래 메뉴 셀들의 구분자다. "셀프라면코너"처럼 원문자가 없는 표기는
# 메뉴 이름 자체이므로 여기 매치하지 않는다.
_CORNER_LABEL_RE = re.compile(r"^[①-⑨]\s*코너$")


# ── 다운로드 ─────────────────────────────────────────────────────────────────

async def _fetch_bookcode() -> str:
    client = await get_http_client()
    resp = await client.get(VIEWER_URL, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    m = _BOOKCODE_RE.search(resp.text)
    if not m:
        raise RuntimeError("bookcode를 viewer 페이지에서 찾지 못함")
    return m.group(1)


async def _fetch_file_url(bookcode: str) -> tuple[str, str]:
    """RawFileList XML에서 (file_name, file_url) 추출."""
    client = await get_http_client()
    resp = await client.post(
        RAWFILE_URL,
        headers={
            "X-Requested-With": "XMLHttpRequest",
            "Referer": VIEWER_URL,
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        },
        data={"key": WEB_KEY, "bookcode": bookcode, "base64": "N"},
    )
    resp.raise_for_status()
    xml = resp.text
    m_url = _FILE_URL_RE.search(xml)
    m_name = _FILE_NAME_RE.search(xml)
    if not m_url or not m_name:
        raise RuntimeError("RawFileList 응답에서 file_url/name을 찾지 못함")
    return m_name.group(1), m_url.group(1)


async def _download_xlsx(url: str) -> bytes:
    client = await get_http_client()
    resp = await client.get(url)
    resp.raise_for_status()
    return resp.content


# ── 파싱 ─────────────────────────────────────────────────────────────────────

def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _extract_time(text: str) -> str | None:
    m = _TIME_RE.search(text)
    return f"{m.group(1)}~{m.group(2)}" if m else None


def _find_date_columns(ws, row: int) -> list[tuple[int, str]]:
    cols = []
    for c in range(1, ws.max_column + 1):
        text = _clean(ws.cell(row=row, column=c).value)
        m = _DATE_HEADER_RE.match(text)
        if m:
            cols.append((c, m.group(2)))
    return cols


def _meal_row_ranges(ws, start: int, end: int) -> list[tuple[int, int, str, str | None]]:
    """식사 헤더의 (시작행, 끝행, 식사타입, 시간) 리스트.

    천원의 아침밥처럼 A열에 두 개의 merged 셀(헤더 + 시간 표기)이 인접해 있는
    경우, 두 영역을 한 식사로 묶는다. 그렇지 않으면 ②코너(R5~R6) 같은 후행 행이
    어떤 식사에도 속하지 않게 된다.
    """
    merged_by_top = {(m.min_row, m.min_col): m for m in ws.merged_cells.ranges}
    result = []
    for r in range(start, end + 1):
        text = _clean(ws.cell(row=r, column=1).value)
        if not _MEAL_HEADER_RE.search(text):
            continue
        mtype = re.sub(r"\s+", " ", _MEAL_HEADER_RE.search(text).group(1))
        time = _extract_time(text)
        if (r, 1) in merged_by_top:
            end_row = merged_by_top[(r, 1)].max_row
        else:
            end_row = r

        # 헤더 직후에 또 다른 A열 merged 셀이 있고 그 텍스트가 시간 표기면
        # 같은 식사의 메타 영역으로 보고 범위를 확장한다.
        next_r = end_row + 1
        if (next_r, 1) in merged_by_top:
            next_text = _clean(ws.cell(row=next_r, column=1).value)
            if _TIME_RE.search(next_text):
                if not time:
                    time = _extract_time(next_text)
                end_row = merged_by_top[(next_r, 1)].max_row

        result.append((r, end_row, mtype, time))
    return result


def _collect_menu(ws, start_row: int, end_row: int, col: int) -> list[str]:
    """day 열의 메뉴 항목들을 모은다.

    "①코너" 같은 코너 라벨 셀은 그 자체로 메뉴가 아니라, 다음 라벨이 나오기
    전까지 뒤따르는 메뉴 셀들의 구분자다. 예전엔 라벨 셀과 메뉴 셀을 각각
    독립된 배열 원소로 그대로 내보내서, 프런트가 "①코너"를 메뉴 이름처럼
    태그로 그리는 결함이 났다(실측: by_day가
    ["①코너", "떡만두국", "②코너", "김치볶음밥/고로케샐러드"] 형태로 왔다).
    라벨을 뒤따르는 메뉴 앞에 붙여 "①코너 떡만두국" 하나의 문자열로 묶는다.
    """
    items = []
    corner_label = None
    label_used = True
    for r in range(start_row, end_row + 1):
        v = _clean(ws.cell(row=r, column=col).value)
        if not v:
            continue
        if _CORNER_LABEL_RE.match(v):
            corner_label = v
            label_used = False
            continue
        items.append(f"{corner_label} {v}" if corner_label else v)
        label_used = True
    # 코너 라벨만 있고 뒤따르는 메뉴가 없는 예외 상황 — 라벨을 그대로 남겨
    # 데이터를 잃지 않는다(실제로는 관측되지 않았지만 지어내지 않기 위한 안전망).
    if not label_used:
        items.append(corner_label)
    return items


def _parse_section(ws, title_row: int, last_row: int, name: str) -> dict:
    date_cols = _find_date_columns(ws, title_row + 1)
    if not date_cols:
        return {"name": name, "meals": []}
    meals = []
    for hr, end_row, mtype, time in _meal_row_ranges(ws, title_row + 2, last_row):
        by_day = {day: _collect_menu(ws, hr, end_row, col) for col, day in date_cols}
        meals.append({"type": mtype, "time": time, "by_day": by_day})
    return {"name": name, "meals": meals}


def _parse_xlsx(content: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    week_start = ws.title  # "5.11"

    tip_row = e_row = None
    for m in ws.merged_cells.ranges:
        if m.min_row == m.max_row and m.min_col == 1 and m.max_col >= 6:
            text = _clean(ws.cell(row=m.min_row, column=1).value)
            if _TITLE_TIP_RE.search(text):
                tip_row = m.min_row
            elif _TITLE_E_RE.search(text):
                e_row = m.min_row

    cafeterias = []
    if tip_row is not None:
        last = (e_row - 1) if e_row is not None else ws.max_row
        cafeterias.append(_parse_section(ws, tip_row, last, "TIP 학생식당"))
    if e_row is not None:
        cafeterias.append(_parse_section(ws, e_row, ws.max_row, "E동 레스토랑"))

    return {"week_start": week_start, "cafeterias": cafeterias}


def _resolve_week_year(week_start: str, now: datetime) -> int:
    """시트명(M.D)과 현재 KST 시각으로 그 주차의 연도를 정한다.

    보통은 받은 시점의 연도가 곧 주차의 연도지만, 12월 말에 시작하는 주차를
    해가 바뀐 뒤 받으면(또는 1월 첫 주차를 12월 말에 미리 받으면) 한 해가
    어긋난다. 프런트가 이 연도로 요일을 계산하므로 여기서 보정한다.
    """
    try:
        month = int(str(week_start).split(".")[0])
    except (TypeError, ValueError):
        return now.year
    if month == 12 and now.month == 1:
        return now.year - 1
    if month == 1 and now.month == 12:
        return now.year + 1
    return now.year


# ── 공개 API ────────────────────────────────────────────────────────────────

async def get_menu(force_refresh: bool = False) -> dict | None:
    """주간 식단표 가져오기 (캐시 우선). force_refresh=True면 캐시 무시."""
    if not force_refresh:
        cached = await get_cached_json(CACHE_KEY)
        if cached:
            return cached

    try:
        bookcode = await _fetch_bookcode()
        file_name, file_url = await _fetch_file_url(bookcode)
        content = await _download_xlsx(file_url)
        parsed = _parse_xlsx(content)
    except Exception:
        logger.exception("학식 식단표 갱신 실패")
        # 실패 시에도 stale 캐시는 살려둠 (만약 있다면) — 아래 None은 cold start만
        cached = await get_cached_json(CACHE_KEY)
        return cached

    now = datetime.now(KST)
    result = {
        **parsed,
        "year": _resolve_week_year(parsed.get("week_start", ""), now),
        "source_file": file_name,
        "fetched_at": now.isoformat(),
    }
    await set_cached_json(CACHE_KEY, result, ttl=CACHE_TTL)
    return result


async def refresh_menu() -> dict | None:
    """APScheduler가 호출하는 강제 갱신 진입점."""
    return await get_menu(force_refresh=True)
